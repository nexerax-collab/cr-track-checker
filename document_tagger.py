import streamlit as st
import os
import io
import json
import PyPDF2
from datetime import datetime
import google.generativeai as genai
# New imports for new file types
import docx
import openpyxl
# Imports for OCR functionality
from PIL import Image
import pytesseract
from pdf2image import convert_from_bytes


# --- App Configuration ---
st.set_page_config(
    page_title="Document Classification & Tagging",
    page_icon="🤖",
    layout="wide"
)

# --- Predefined PLM/PDM Document Types ---
PLM_DOCUMENT_TYPES = [
    "Technical Specification",
    "Requirements Document (System/Software)",
    "Bill of Materials (BOM)",
    "Engineering Change Request (ECR)",
    "Engineering Change Order (ECO)",
    "CAD Drawing (2D/3D)",
    "Failure Mode and Effects Analysis (FMEA)",
    "Test Plan / Test Case",
    "Test Report / Validation Report",
    "User Manual / Operator Guide",
    "Manufacturing Process Plan",
    "Quality Inspection Report",
    "Supplier Qualification Document",
    "Project Plan",
    "Risk Analysis Report",
    "Other"  # Fallback category
]

# --- Session State Initialization ---
def init_session_state():
    """Initializes session state variables if they don't exist."""
    defaults = {
        'processed_documents': {},
        'google_api_key': None,
        'edit_mode': {} # Tracks which document is in edit mode, e.g., {'filename.pdf': True}
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value
init_session_state()

# --- Helper Functions for Text Extraction ---

def perform_ocr(file_bytes: bytes) -> str:
    """Performs OCR on a PDF if it's image-based."""
    st.info("Standard text extraction failed. Attempting OCR on the document. This might be slow for large files.")
    try:
        images = convert_from_bytes(file_bytes)
        ocr_text = ""
        for i, image in enumerate(images):
            st.write(f"Processing page {i+1} with OCR...")
            ocr_text += pytesseract.image_to_string(image) + "\n"
        return ocr_text
    except Exception as e:
        st.error(f"OCR processing failed. Please ensure Tesseract is installed and configured correctly. Error: {e}")
        return ""


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extracts text from a PDF, falling back to OCR if needed."""
    text = ""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        for page in pdf_reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        
        if len(text.strip()) < 50: # Arbitrary threshold to detect scanned PDFs
            ocr_text = perform_ocr(file_bytes)
            return ocr_text if ocr_text else text
        return text
        
    except Exception as e:
        st.error(f"Error reading PDF file: {e}. Attempting OCR as a fallback.")
        return perform_ocr(file_bytes)

def extract_text_from_docx(file_bytes: bytes) -> str:
    """Extracts text from a .docx file."""
    try:
        document = docx.Document(io.BytesIO(file_bytes))
        return "\n".join([para.text for para in document.paragraphs])
    except Exception as e:
        st.error(f"Error reading Word document: {e}")
        return ""

def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extracts text from all cells in an .xlsx file."""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
        text = ""
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value:
                        text += str(cell.value) + " "
                text += "\n"
        return text
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return ""
        
def extract_text_from_txt(file_bytes: bytes) -> str:
    """Extracts text from a .txt file."""
    try:
        return file_bytes.decode('utf-8')
    except Exception as e:
        st.error(f"Error reading text file: {e}")
        return ""


def get_gemini_response(api_key: str, text_content: str) -> dict | None:
    """Analyzes document text using the Gemini API and returns a structured response."""
    if not text_content:
        st.warning("Document appears to be empty or unreadable. Could not extract text for analysis.")
        return None
        
    try:
        genai.configure(api_key=api_key)
        
        generation_config = {"temperature": 0.2, "response_mime_type": "application/json"}
        safety_settings = [
            {"category": "HARM_CATEGORY_HARASSMENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
            {"category": "HARM_CATEGORY_HATE_SPEECH", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
            {"category": "HARM_CATEGORY_SEXUALLY_EXPLICIT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
            {"category": "HARM_CATEGORY_DANGEROUS_CONTENT", "threshold": "BLOCK_MEDIUM_AND_ABOVE"},
        ]
        
        schema = {
            "type": "object",
            "properties": {
                "category": {"type": "string", "description": "The most likely document type."},
                "confidence_score": {"type": "integer", "description": "A score from 0 to 100 representing certainty."},
                "tags": {"type": "array", "items": {"type": "string"}, "description": "A list of 5-7 relevant keywords."},
                "reasoning": {"type": "string", "description": "A brief justification for the chosen category and confidence."}
            },
            "required": ["category", "confidence_score", "tags", "reasoning"]
        }

        model = genai.GenerativeModel(model_name="gemini-1.5-flash", generation_config=generation_config, safety_settings=safety_settings)
        
        prompt = f"""
        You are an expert in Product Lifecycle Management (PLM) and document control.
        Your task is to analyze the provided technical document text and classify it.
        
        Based on the text content, perform the following actions and provide your response ONLY as a valid JSON object matching the defined schema:
        1.  **category**: From the following list, choose the single most likely document type: {json.dumps(PLM_DOCUMENT_TYPES)}. If none fit, use 'Other'.
        2.  **confidence_score**: Provide an integer score from 0 to 100 indicating how confident you are.
        3.  **tags**: Generate a list of 5 to 7 relevant keywords.
        4.  **reasoning**: Briefly explain your choice in one or two sentences.
        
        Here is the document text to analyze (first 8000 characters):
        ---
        {text_content[:8000]}
        ---
        """

        response = model.generate_content([prompt, schema])
        cleaned_response_text = response.text.strip().replace("```json", "").replace("```", "").strip()
        result_json = json.loads(cleaned_response_text)
        return result_json

    except genai.types.generation_types.StopCandidateException as e:
        st.error(f"AI analysis was stopped. This can happen if the content is flagged by safety filters. Details: {e}")
        return None
    except Exception as e:
        if "API_KEY_INVALID" in str(e):
            st.error("The provided Google AI API Key is invalid. Please check and re-enter it in the sidebar.")
        else:
            st.error(f"An error occurred with the AI analysis: {e}")
        return None

# --- UI Rendering Functions ---

def render_upload_page():
    """Renders the main page for uploading and analyzing documents."""
    st.title("📄 AI-Powered Document Classification & Tagging")
    
    st.markdown("""
    ### Welcome! Make Your Document Management Smarter.
    This tool, powered by Google's Gemini AI, helps you automatically process and organize your technical documents.
    
    **Why is this helpful?**
    - **Saves Time:** Instantly categorize documents that would otherwise require manual review.
    - **Ensures Consistency:** Uses a predefined set of categories, standardizing your document management.
    - **Improves Findability:** Generates relevant tags, making it easier to search and retrieve documents later.
    - **Handles Multiple Formats:** Processes PDF, Word (.docx), Excel (.xlsx), and Text (.txt) files.
    - **Handles Scanned Documents:** Includes Optical Character Recognition (OCR) to read text from image-based PDFs.
    
    **How to get started:**
    1.  **Enter API Key:** Provide your Google AI API key in the sidebar.
    2.  **(One-Time OCR Setup):** For scanned documents, you must install Tesseract OCR on your system. See instructions [here](https://tesseract-ocr.github.io/tessdoc/Installation.html). You also need to install `poppler`.
    3.  **Upload:** Drag and drop your documents below.
    4.  **Analyze:** Click "Start Analysis" and let the AI do the work.
    """)
    
    # Updated file uploader to accept new types
    uploaded_files = st.file_uploader(
        "Upload PDF, Word, Excel, or Text files", 
        type=['pdf', 'docx', 'xlsx', 'txt'], 
        accept_multiple_files=True
    )
    
    if st.button("Start Analysis", disabled=(not uploaded_files or not st.session_state.google_api_key)):
        if not uploaded_files: st.warning("Please upload at least one file.")
        if not st.session_state.google_api_key: st.warning("Please enter your Google AI API key in the sidebar.")
        if uploaded_files and st.session_state.google_api_key:
            with st.spinner("Analyzing documents... This may take a moment."):
                for file in uploaded_files:
                    if file.name in st.session_state.processed_documents:
                        st.info(f"'{file.name}' has already been processed. Skipping.")
                        continue
                    
                    st.write(f"--- \n**Processing: {file.name}**")
                    file_bytes = file.getvalue()
                    
                    # Determine file type and call appropriate extraction function
                    text = ""
                    file_extension = os.path.splitext(file.name)[1].lower()
                    
                    if file_extension == ".pdf":
                        text = extract_text_from_pdf(file_bytes)
                    elif file_extension == ".docx":
                        text = extract_text_from_docx(file_bytes)
                    elif file_extension == ".xlsx":
                        text = extract_text_from_xlsx(file_bytes)
                    elif file_extension == ".txt":
                        text = extract_text_from_txt(file_bytes)

                    if text:
                        ai_result = get_gemini_response(st.session_state.google_api_key, text)
                        if ai_result:
                            confidence = ai_result.get("confidence_score", 0)
                            status = "Auto-Classified" if confidence >= 50 else "Needs Verification"
                            st.session_state.processed_documents[file.name] = {
                                "filename": file.name, "category": ai_result.get("category", "N/A"),
                                "confidence": confidence, "tags": ai_result.get("tags", []),
                                "reasoning": ai_result.get("reasoning", "No reasoning provided."),
                                "status": status, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                            st.success(f"'{file.name}' analyzed. Status: {status}")
                        else:
                            st.session_state.processed_documents[file.name] = {
                                "filename": file.name, "category": "Error", "confidence": 0, "tags": [],
                                "reasoning": "AI analysis failed. Please review manually.", "status": "Error",
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            }
                    else:
                        st.session_state.processed_documents[file.name] = {
                            "filename": file.name, "category": "Unreadable", "confidence": 0, "tags": [],
                            "reasoning": f"Could not extract text from the {file_extension} file. It might be empty, corrupted, or require special handling (e.g., OCR).",
                            "status": "Error",
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
            st.success("Analysis complete! Check the 'Classification Results' page for details.")

def render_results_page():
    """Renders the page displaying the classification results."""
    st.title("📊 Classification Results")
    
    if not st.session_state.processed_documents:
        st.info("No documents have been processed yet. Please go to the 'Upload Document' page to begin.")
        return

    st.markdown("Review the analysis results below. For documents marked 'Needs Verification', you can manually edit the category and tags.")
    
    for filename, details in st.session_state.processed_documents.items():
        is_editing = st.session_state.edit_mode.get(filename, False)

        status_color = "green" if details['status'] == "Auto-Classified" else \
                       "blue" if details['status'] == "Manually Verified" else \
                       "orange" if details['status'] == "Needs Verification" else "red"
        
        with st.expander(f"**{details['filename']}** | Status: :{status_color}[{details['status']}]", expanded=is_editing):
            if is_editing:
                st.write("**Editing Classification Details**")
                try:
                    current_cat_index = PLM_DOCUMENT_TYPES.index(details['category'])
                except ValueError:
                    if details['category'] not in PLM_DOCUMENT_TYPES: PLM_DOCUMENT_TYPES.append(details['category'])
                    current_cat_index = PLM_DOCUMENT_TYPES.index(details['category'])
                
                new_category = st.selectbox("Document Category", options=PLM_DOCUMENT_TYPES, index=current_cat_index, key=f"cat_edit_{filename}")
                new_tags = st.text_area("Tags (comma-separated)", value=", ".join(details['tags']), key=f"tags_edit_{filename}")

                col1, col2, _ = st.columns([1, 1, 4])
                if col1.button("Save Changes", key=f"save_{filename}", type="primary"):
                    st.session_state.processed_documents[filename]['category'] = new_category
                    st.session_state.processed_documents[filename]['tags'] = [tag.strip() for tag in new_tags.split(",") if tag.strip()]
                    st.session_state.processed_documents[filename]['status'] = "Manually Verified"
                    st.session_state.edit_mode[filename] = False
                    st.rerun()
                if col2.button("Cancel", key=f"cancel_{filename}"):
                    st.session_state.edit_mode[filename] = False
                    st.rerun()
            else:
                col1, col2 = st.columns([1, 1])
                with col1:
                    st.metric("Suggested Category", details['category'])
                    st.metric("Confidence Score", f"{details['confidence']}%" if details['status'] != "Manually Verified" else "N/A")
                with col2:
                    st.write("**Suggested Tags**")
                    st.multiselect("Tags", options=details['tags'], default=details['tags'], disabled=True, label_visibility="collapsed", key=f"tags_view_{filename}")
                    st.write(f"**Processed:** {details['timestamp']}")
                
                st.write("**AI Reasoning**")
                st.info(details['reasoning'])

                if st.button("Edit", key=f"edit_{filename}"):
                    st.session_state.edit_mode[filename] = True
                    st.rerun()

# --- Main App Logic ---

with st.sidebar:
    st.header("Configuration")
    api_key_input = st.text_input("Enter your Google AI API Key", type="password", key="google_api_key_input")
    if api_key_input:
        st.session_state.google_api_key = api_key_input
        st.success("API Key set successfully!")

    st.sidebar.title("Navigation")
    app_mode = st.sidebar.radio("Choose a page:", ["Upload Document", "Classification Results"])

if app_mode == "Upload Document":
    render_upload_page()
elif app_mode == "Classification Results":
    render_results_page()
